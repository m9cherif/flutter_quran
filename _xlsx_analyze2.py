import openpyxl, json, sys

page = "553"
wb = openpyxl.load_workbook(rf'G:\flutter_quran_data\annotation\a{page}.xlsx', data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {headers}")
    if sheet_name == "Mots" and ws.max_row > 1:
        # Count unique in selected columns
        for i, h in enumerate(headers):
            col_vals = [ws.cell(r, i+1).value for r in range(2, ws.max_row + 1)]
            unique = set(col_vals)
            print(f"  '{h}': {len(unique)} unique, types={set(type(v).__name__ for v in unique)}")

# Timeline analysis
d = json.load(open(r'G:\flutter_quran_data\timeline\page553.json', 'r', encoding='utf-8'))
events = d['events']
timeline_ids = set(e['word_id'] for e in events)
print(f"\nTimeline word_ids: {min(timeline_ids)}-{max(timeline_ids)}, count={len(timeline_ids)}")
from collections import Counter
cnt = Counter(e['word_id'] for e in events)
repeated = [(wid, c) for wid, c in cnt.items() if c > 1]
print(f"Repeated word_ids: {repeated}")

# Check what aya_no values look like in XLSX
ws = wb['Mots']
aya_values = set()
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 9).value  # aya_no
    aya_values.add(v)
print(f"\naya_no values: {sorted(aya_values)[:20]}...")
print(f"Total unique aya_no: {len(aya_values)}")
