import openpyxl, json

page = "553"
wb = openpyxl.load_workbook(rf'G:\flutter_quran_data\annotation\a{page}.xlsx', data_only=True)
print(f"Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Headers: {headers}")
    if sheet_name == "Mots" and ws.max_row > 1:
        print(f"First 5 rows:")
        for r in range(2, min(7, ws.max_row + 1)):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            print(f"  {vals}")
        print(f"Last 5 rows:")
        for r in range(max(2, ws.max_row - 4), ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            print(f"  {vals}")
        # Count unique values in each column
        for i, h in enumerate(headers):
            col_vals = [ws.cell(r, i+1).value for r in range(2, ws.max_row + 1)]
            unique = set(col_vals)
            print(f"  Column '{h}': {len(unique)} unique, sample={list(unique)[:5]}")

# Also check which word_ids appear in timeline
d = json.load(open(r'G:\flutter_quran_data\timeline\page553.json', 'r', encoding='utf-8'))
events = d['events']
timeline_ids = set(e['word_id'] for e in events)
print(f"\nTimeline word_ids: {sorted(timeline_ids)}")
# Word_ids that appear twice
from collections import Counter
cnt = Counter(e['word_id'] for e in events)
repeated = [wid for wid, c in cnt.items() if c > 1]
print(f"Repeated word_ids: {repeated}")
